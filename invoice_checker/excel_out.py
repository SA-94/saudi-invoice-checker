# -*- coding: utf-8 -*-
"""كتابة نتائج الفحص في ملف إكسل جاهز للاستخدام."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import validate
from .validate import MISSING, OK, REVIEW

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
STATUS_FILL = {
    OK: PatternFill("solid", fgColor="C6EFCE"),
    REVIEW: PatternFill("solid", fgColor="FFEB9C"),
    MISSING: PatternFill("solid", fgColor="FFC7CE"),
}
STATUS_FONT = {
    OK: Font(color="006100", bold=True),
    REVIEW: Font(color="9C5700", bold=True),
    MISSING: Font(color="9C0006", bold=True),
}
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = "#,##0.00"
TEXT_FMT = "@"
DATE_FMT = "yyyy-mm-dd"

# ورقة البيانات: الأعمدة الستة المطلوبة فقط.
# الحالة والتنبيهات واسم الملف في ورقتَي «الفحص التفصيلي» و«التنبيهات».
# (العنوان، مفتاح السجل، العرض، التنسيق)
MAIN_COLUMNS = [
    ("تاريخ الفاتورة", "invoice_date", 16, DATE_FMT),
    ("اسم المورد", "seller_name", 30, None),
    ("الرقم الضريبي للمورد", "seller_vat", 22, TEXT_FMT),
    ("المبلغ قبل الضريبة", "net_amount", 18, MONEY_FMT),
    ("الضريبة", "vat_amount", 14, MONEY_FMT),
    ("المبلغ شامل الضريبة", "total_amount", 20, MONEY_FMT),
]
SUM_KEYS = ("net_amount", "vat_amount", "total_amount")


def _style_header(ws, headers):
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    if len(headers) >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_data(wb, records):
    ws = wb.active
    ws.title = "البيانات"
    _style_header(ws, [c[0] for c in MAIN_COLUMNS])

    for col, (_, _, width, _) in enumerate(MAIN_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row, rec in enumerate(records, start=2):
        for col, (_, key, _, fmt) in enumerate(MAIN_COLUMNS, 1):
            value = rec.get(key)

            # المبالغ أرقام حقيقية عشان تنجمع، والتاريخ تاريخ حقيقي عشان يتفرز
            if fmt == MONEY_FMT and value is not None:
                value = float(value)
            elif fmt == DATE_FMT:
                value = validate.parse_date(value) or validate.clean_text(value)
            elif fmt == TEXT_FMT and value is not None:
                value = str(value)

            cell = ws.cell(row=row, column=col, value=value if value is not None else "—")
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="right" if fmt is None else "center", vertical="center"
            )
            if fmt and value is not None:
                cell.number_format = fmt
            # نلوّن الصف حسب حالة الفحص عشان تبان المشاكل بالعين
            if rec.get("status") != OK:
                cell.fill = STATUS_FILL.get(rec.get("status"), PatternFill())

    # صف المجموع
    if records:
        total_row = len(records) + 2
        label = ws.cell(row=total_row, column=2, value="الإجمالي")
        label.font = Font(bold=True)
        label.alignment = Alignment(horizontal="right", vertical="center")
        for col, (_, key, _, _) in enumerate(MAIN_COLUMNS, 1):
            if key not in SUM_KEYS:
                continue
            letter = get_column_letter(col)
            cell = ws.cell(
                row=total_row, column=col, value=f"=SUM({letter}2:{letter}{total_row - 1})"
            )
            cell.number_format, cell.font = MONEY_FMT, Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF1F5")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return ws


def _sheet_checks(wb, records):
    all_checks = []
    for rec in records:
        for name in rec.get("checks", {}):
            if name not in all_checks:
                all_checks.append(name)
    if not all_checks:
        return

    ws = wb.create_sheet("الفحص التفصيلي")
    _style_header(ws, ["م", "اسم الملف", "الحالة"] + all_checks)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    for col in range(4, 4 + len(all_checks)):
        ws.column_dimensions[get_column_letter(col)].width = 15

    for row, rec in enumerate(records, start=2):
        ws.cell(row=row, column=1, value=row - 1).border = BORDER
        ws.cell(row=row, column=2, value=rec.get("file_name")).border = BORDER
        cell = ws.cell(row=row, column=3, value=rec.get("status"))
        cell.fill = STATUS_FILL.get(rec.get("status"), PatternFill())
        cell.font = STATUS_FONT.get(rec.get("status"), Font())
        cell.border = BORDER

        checks = rec.get("checks", {})
        for i, name in enumerate(all_checks):
            state = checks.get(name)
            cell = ws.cell(
                row=row, column=4 + i,
                value="✔" if state is True else ("✘" if state is False else "—"),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if state is True:
                cell.font = Font(color="006100", bold=True, size=13)
            elif state is False:
                cell.font = Font(color="9C0006", bold=True, size=13)
                cell.fill = STATUS_FILL[MISSING]


def _sheet_alerts(wb, records):
    flagged = [r for r in records if r.get("status") != OK]
    ws = wb.create_sheet("التنبيهات")
    _style_header(ws, ["م", "اسم الملف", "الحالة", "التنبيه"])
    for width, letter in ((5, "A"), (36, "B"), (14, "C"), (70, "D")):
        ws.column_dimensions[letter].width = width

    if not flagged:
        ws.cell(row=2, column=2, value="ما فيه أي تنبيه — كل الفواتير سليمة ✔").font = Font(
            bold=True, color="006100", size=12
        )
        return

    row = 2
    for rec in flagged:
        alerts = rec.get("alerts") or ["—"]
        for alert in alerts:
            ws.cell(row=row, column=1, value=row - 1).border = BORDER
            ws.cell(row=row, column=2, value=rec.get("file_name")).border = BORDER
            cell = ws.cell(row=row, column=3, value=rec.get("status"))
            cell.fill = STATUS_FILL.get(rec.get("status"), PatternFill())
            cell.font = STATUS_FONT.get(rec.get("status"), Font())
            cell.border = BORDER
            c = ws.cell(row=row, column=4, value=alert)
            c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            c.border = BORDER
            row += 1


def write(records, path):
    """يكتب كل النتائج في ملف إكسل بثلاث أوراق."""
    wb = Workbook()
    _sheet_data(wb, records)
    _sheet_checks(wb, records)
    _sheet_alerts(wb, records)
    wb.save(path)
    return path
